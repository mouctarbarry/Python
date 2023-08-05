import openpyxl as xl

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    print(cell.value)
